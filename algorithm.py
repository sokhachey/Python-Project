import os
import openpyxl
import speedtest
import math
import psutil
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter.scrolledtext import ScrolledText
import threading
import platform
import socket
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import getpass
import wmi  
import subprocess
from tkinter import messagebox
from openpyxl.styles import Font, Alignment, PatternFill
from tkinter import ttk


def get_system_data():
    # Initialize WMI for hardware information
    w = wmi.WMI()

    # Collect basic system info
    cores = os.cpu_count()
    ram = f"{math.ceil(psutil.virtual_memory().total / (1024 ** 3))}GB"
    disk = f"{math.ceil(psutil.disk_usage('/').total / (1024 ** 3))}GB"

    specification = f"{cores} Cores, {ram}, {disk}"
    username = getpass.getuser()

    # Collect hardware details
    try:
        system_info = w.Win32_ComputerSystem()[0]
        serial_number = w.Win32_BIOS()[0].SerialNumber
        system_model = system_info.Model
        system_manufacturer = system_info.Manufacturer
    except Exception as e:
        serial_number = system_model = system_manufacturer = "N/A"

    # Get the Hostname instead of Asset Tag
    hostname = socket.gethostname()

    # Return data as a dictionary with an added "Number" key
    return {
        "Specification": specification,
        "Username (Hostname)": f"{username}",
        "Serial Number": serial_number,
        "System Model": system_model,
        "System Manufacturer": system_manufacturer,
        "Asset Tag": hostname,
    }


def write_data_to_excel(data, file_name="system_data.xlsx"):
    # Check if the file exists
    if not os.path.exists(file_name):
        # Create a new workbook and add headers if the file doesn't exist
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "System Specifications"
        header = ["No."] + list(data.keys())

        # Apply header styles
        header_fill = PatternFill(start_color="00B4F0", end_color="00B4F0", fill_type="solid")
        header_font = Font(bold=True)
        alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header_text in enumerate(header, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment

        workbook.save(file_name)

    # Load existing workbook and find the next empty row
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    next_row = sheet.max_row + 1

    # Write data
    sheet.cell(row=next_row, column=1, value=next_row - 1)  # No. column
    for col_idx, value in enumerate(data.values(), start=2):
        cell = sheet.cell(row=next_row, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save the workbook
    workbook.save(file_name)
    print(f"Data written to {file_name}")


if __name__ == "__main__":
    # Fetch system data
    system_data = get_system_data()

    # Write to Excel
    write_data_to_excel(system_data)



def get_wmi_info(wmi_class, attributes):
    """
    Helper function to get WMI information safely.
    """
    w = wmi.WMI()
    try:
        info = w.query(f"SELECT {', '.join(attributes)} FROM {wmi_class}")
        return info[0] if info else None
    except Exception as e:
        print(f"Error retrieving WMI info from {wmi_class}: {e}")
        return None


def get_hardware_info():
    """
    Retrieve detailed hardware information using WMI.
    """
    hardware_info = {}

    # Get computer system information
    system_info = get_wmi_info("Win32_ComputerSystem", ["Manufacturer", "Model", "SystemType", "Name"])
    if system_info:
        hardware_info['System Manufacturer'] = getattr(system_info, 'Manufacturer', 'N/A')
        hardware_info['System Model'] = getattr(system_info, 'Model', 'N/A')
        hardware_info['System Type'] = getattr(system_info, 'SystemType', 'N/A')

    # Get BIOS serial number
    bios_info = get_wmi_info("Win32_BIOS", ["SerialNumber"])
    hardware_info['Serial Number'] = getattr(bios_info, 'SerialNumber', 'N/A')

    return hardware_info


def get_os_info():
    """
    Retrieve operating system information.
    """
    os_info = {
        "Operating System": f"{platform.system()} {platform.release()}",
        "OS Version": platform.version(),
        "OS Architecture": platform.architecture()[0],
        "Windows Edition": get_windows_edition(),
        "Hostname": socket.gethostname(),
        "Username": getpass.getuser(),
    }
    return os_info


def get_processor_info():
    """
    Retrieve processor information using WMI and platform.
    """
    processor_info = {}

    # Get processor details
    cpu_info = get_wmi_info("Win32_Processor", ["Name", "MaxClockSpeed", "NumberOfCores", "NumberOfLogicalProcessors"])
    if cpu_info:
        processor_info['Processor Name'] = getattr(cpu_info, 'Name', 'N/A')
        processor_info['Max Clock Speed (MHz)'] = getattr(cpu_info, 'MaxClockSpeed', 'N/A')
        processor_info['Physical Cores'] = getattr(cpu_info, 'NumberOfCores', 'N/A')
        processor_info['Logical Cores'] = getattr(cpu_info, 'NumberOfLogicalProcessors', 'N/A')

    return processor_info


def get_memory_info():
    """
    Retrieve memory (RAM) details using psutil.
    """
    memory = psutil.virtual_memory()
    memory_info = {
        "Total RAM (GB)": math.ceil(memory.total / (1024 ** 3)),
        "Used RAM (GB)": math.ceil(memory.used / (1024 ** 3)),
        "Free RAM (GB)": math.ceil(memory.available / (1024 ** 3)),
    }
    return memory_info


def get_disk_info():
    """
    Retrieve disk storage details using psutil.
    """
    disk = psutil.disk_usage('/')
    disk_info = {
        "Total Disk (GB)": math.ceil(disk.total / (1024 ** 3)),
        "Used Disk (GB)": math.ceil(disk.used / (1024 ** 3)),
        "Free Disk (GB)": math.ceil(disk.free / (1024 ** 3)),
    }
    return disk_info


def get_network_info():
    """
    Retrieve network information (IP and MAC addresses).
    """
    try:
        hostname = socket.gethostname()
        ip_address = socket.gethostbyname(hostname)
        mac_address = ':'.join(['{:02x}'.format((psutil.net_if_addrs()['Ethernet'][0].address[i])) for i in range(6)])
    except Exception:
        ip_address = mac_address = "N/A"

    return {
        "IP Address": ip_address,
        "MAC Address": mac_address
    }

def get_windows_edition():
    """
    Retrieve the Windows Edition using WMI.
    """
    os_info = get_wmi_info("Win32_OperatingSystem", ["Caption"])
    return getattr(os_info, 'Caption', 'N/A')


def gather_system_info():
    """
    Gather and combine all system information.
    """
    system_info = {}
    system_info.update(get_os_info())
    system_info.update(get_hardware_info())
    system_info.update(get_processor_info())
    system_info.update(get_memory_info())
    system_info.update(get_disk_info())
    system_info.update(get_network_info())

    return system_info


def print_system_info(info):
    """
    Nicely format and print system information.
    """
    print("\nSystem Information:")
    for key, value in info.items():
        print(f"{key}: {value}")


# Shutdown and Restart
def shutdown():
    os.system("shutdown /s /t 1")

def restart():
    os.system("shutdown /r /t 1")


# Function to Test Wi-Fi Speed
def test_wifi_speed():
    wifi_speed = None
    try:
        st = speedtest.Speedtest()
        st.get_best_server()
        download_speed = st.download() / 1_000_000  
        upload_speed = st.upload() / 1_000_000  
        ping = st.results.ping
        wifi_speed = f"Download: {download_speed:.2f} Mbps\nUpload: {upload_speed:.2f} Mbps\nPing: {ping} ms"
    except Exception as e:
        wifi_speed = f"Error testing speed: {e}"

    return wifi_speed


# Display Wi-Fi Speed Test Results
def display_wifi_speed():
    speed = test_wifi_speed()
    actions_tab.delete(1.0, tk.END) 
    actions_tab.insert(tk.END, speed)


# Retrieve Wi-Fi Passwords (Windows Only)
def get_wifi_passwords():
    wifi_passwords = []
    
    try:
        profiles = subprocess.check_output("netsh wlan show profiles").decode("utf-8", errors="backslashreplace").split('\n')
        profile_names = [i.split(":")[1][1:-1] for i in profiles if "All User Profile" in i]

        for profile in profile_names:
            try:
                password_info = subprocess.check_output(f'netsh wlan show profile "{profile}" key=clear', shell=True).decode("utf-8", errors="backslashreplace")
                password = [i.split(":")[1][1:-1] for i in password_info.split("\n") if "Key Content" in i]
                if password:
                    wifi_passwords.append((profile, password[0]))
            except subprocess.CalledProcessError:
                wifi_passwords.append((profile, "No password set"))
    
        if not wifi_passwords:
            return "No Wi-Fi profiles found."
        return wifi_passwords
    except Exception as e:
        return f"Error retrieving Wi-Fi passwords: {e}"


# Update Real-Time Graphs
def update_graphs():
    global cpu_data, memory_data, disk_data
    cpu_data.append(psutil.cpu_percent())
    memory_data.append(psutil.virtual_memory().percent)
    disk_data.append(psutil.disk_usage('/').percent)
    
    if len(cpu_data) > 50:
        cpu_data.pop(0)
        memory_data.pop(0)
        disk_data.pop(0)

    cpu_line.set_ydata(cpu_data)
    mem_line.set_ydata(memory_data)
    disk_line.set_ydata(disk_data)

    canvas.draw()
    root.after(1000, update_graphs)


# Alert for High CPU Usage
def check_alerts():
    cpu_usage = psutil.cpu_percent()
    memory_usage = psutil.virtual_memory().percent
    disk_usage = psutil.disk_usage('/').percent
    
    if cpu_usage > 80:
        messagebox.showwarning("Alert", f"High CPU Usage: {cpu_usage}%")
    if memory_usage > 85:
        messagebox.showwarning("Alert", f"High Memory Usage: {memory_usage}%")
    if disk_usage > 90:
        messagebox.showwarning("Alert", f"High Disk Usage: {disk_usage}%")
    
    root.after(5000, check_alerts)


# GUI Setup
root = tk.Tk()
root.title("Enhanced System Monitoring Tool")
root.geometry("900x650")
root.configure(bg="#f0f0f0")  

# Set up style for widgets
style = ttk.Style()
style.configure("TButton", font=("Helvetica", 12), padding=6)
style.configure("TLabel", font=("Helvetica", 14, "bold"))
style.configure("TNotebook.Tab", font=("Helvetica", 12))

notebook = ttk.Notebook(root)
notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

# Tab 1: Real-Time Monitoring
monitor_tab = ttk.Frame(notebook)
notebook.add(monitor_tab, text="Real-Time Monitoring")

fig, ax = plt.subplots(figsize=(5, 3))
cpu_data, memory_data, disk_data = [0] * 50, [0] * 50, [0] * 50

cpu_line, = ax.plot(cpu_data, label="CPU Usage (%)")
mem_line, = ax.plot(memory_data, label="Memory Usage (%)")
disk_line, = ax.plot(disk_data, label="Disk Usage (%)")

ax.set_ylim(0, 100)
ax.set_title("Real-Time System Resource Monitoring", fontsize=16, fontweight='bold')
ax.legend(loc="upper right")
canvas = FigureCanvasTkAgg(fig, master=monitor_tab)
canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)


# Tab 2: System Info (Grid Layout)
info_tab = ttk.Frame(notebook)
notebook.add(info_tab, text="System Info")

# Helper function to add data in a label grid
def add_label_value(frame, row, label_text, value_text):
    label = ttk.Label(frame, text=label_text, font=("Helvetica", 12, "bold"))
    label.grid(row=row, column=0, sticky="w", padx=10, pady=5)
    value = ttk.Label(frame, text=value_text, font=("Helvetica", 12))
    value.grid(row=row, column=1, sticky="w", padx=10, pady=5)


# Function to refresh and display info in grouped layout
def refresh_info_grid():
    system_info = gather_system_info()

    # Group 1: Operating System Details
    os_label = ttk.Label(info_tab, text="Operating System Details", font=("Helvetica", 14, "bold"))
    os_label.grid(row=0, column=0, columnspan=2, sticky="w", padx=10, pady=10)

    add_label_value(info_tab, 1, "Operating System", system_info.get("Operating System", "N/A"))
    add_label_value(info_tab, 2, "OS Version", system_info.get("OS Version", "N/A"))
    add_label_value(info_tab, 3, "OS Architecture", system_info.get("OS Architecture", "N/A"))
    add_label_value(info_tab, 4, "Windows Edition", system_info.get("Windows Edition", "N/A"))  
    add_label_value(info_tab, 5, "Hostname", system_info.get("Hostname", "N/A"))
    add_label_value(info_tab, 6, "Username", system_info.get("Username", "N/A"))

    # Group 2: Processor Details
    processor_label = ttk.Label(info_tab, text="Processor Details", font=("Helvetica", 14, "bold"))
    processor_label.grid(row=7, column=0, columnspan=2, sticky="w", padx=10, pady=10)

    add_label_value(info_tab, 8, "Processor Name", system_info.get("Processor Name", "N/A"))
    add_label_value(info_tab, 9, "Max Clock Speed", f"{system_info.get('Max Clock Speed (MHz)', 'N/A')} MHz")
    add_label_value(info_tab, 10, "Physical Cores", system_info.get("Physical Cores", "N/A"))
    add_label_value(info_tab, 11, "Logical Cores", system_info.get("Logical Cores", "N/A"))

    # Group 3: Memory Details
    memory_label = ttk.Label(info_tab, text="Memory (RAM) Details", font=("Helvetica", 14, "bold"))
    memory_label.grid(row=12, column=0, columnspan=2, sticky="w", padx=10, pady=10)

    add_label_value(info_tab, 13, "Total RAM", f"{system_info.get('Total RAM (GB)', 'N/A')} GB")
    add_label_value(info_tab, 14, "Used RAM", f"{system_info.get('Used RAM (GB)', 'N/A')} GB")
    add_label_value(info_tab, 15, "Free RAM", f"{system_info.get('Free RAM (GB)', 'N/A')} GB")

    # Group 4: Disk Details
    disk_label = ttk.Label(info_tab, text="Disk Storage Details", font=("Helvetica", 14, "bold"))
    disk_label.grid(row=16, column=0, columnspan=2, sticky="w", padx=10, pady=10)

    add_label_value(info_tab, 17, "Total Disk", f"{system_info.get('Total Disk (GB)', 'N/A')} GB")
    add_label_value(info_tab, 18, "Used Disk", f"{system_info.get('Used Disk (GB)', 'N/A')} GB")
    add_label_value(info_tab, 19, "Free Disk", f"{system_info.get('Free Disk (GB)', 'N/A')} GB")


# Populate the grid initially
refresh_info_grid()

# Tab 3: Wi-Fi & File Management
wifi_file_tab = ttk.Frame(notebook)
notebook.add(wifi_file_tab, text="Wi-Fi speed & password")

# Add a frame for displaying Wi-Fi speed and passwords
wifi_info_frame = ttk.Frame(wifi_file_tab)
wifi_info_frame.pack(fill=tk.BOTH, padx=10, pady=10, expand=True)

# Label for Wi-Fi Speed
wifi_speed_label = ttk.Label(wifi_info_frame, text="Wi-Fi Speed", font=("Helvetica", 14, "bold"))
wifi_speed_label.pack(anchor=tk.W, padx=10, pady=5)

# Label to display the Wi-Fi speed results
wifi_speed_result_label = ttk.Label(wifi_info_frame, text="Fetching speed...", font=("Helvetica", 12), wraplength=500)
wifi_speed_result_label.pack(anchor=tk.W, padx=10, pady=5)

# Label for Wi-Fi Passwords
wifi_password_label = ttk.Label(wifi_info_frame, text="Wi-Fi Passwords", font=("Helvetica", 14, "bold"))
wifi_password_label.pack(anchor=tk.W, padx=10, pady=5)

# Label to display the Wi-Fi profiles and passwords
wifi_passwords_result_label = ttk.Label(wifi_info_frame, text="Fetching passwords...", font=("Helvetica", 12), wraplength=500)
wifi_passwords_result_label.pack(anchor=tk.W, padx=10, pady=5)

# Function to Test Wi-Fi Speed
def display_wifi_speed():
    def update_ui(speed):
        wifi_speed_result_label.config(text=f"Wi-Fi Speed: {speed}")
    
    # Run the speed test in a background thread
    def run_speed_test():
        speed = test_wifi_speed()  
        root.after(0, update_ui, speed) 
    
    # Start the speed test in a background thread
    threading.Thread(target=run_speed_test).start()

# Function to Retrieve Wi-Fi Passwords
def display_wifi_passwords():
    def update_ui(passwords):
        if isinstance(passwords, list):
            passwords_text = ""
            for profile, password in passwords:
                passwords_text += f"Profile: {profile}\nPassword: {password}\n\n"
            wifi_passwords_result_label.config(text=passwords_text)
        else:
            wifi_passwords_result_label.config(text=passwords) 
    # Run the Wi-Fi password retrieval in a background thread
    def run_wifi_passwords():
        passwords = get_wifi_passwords() 
        root.after(0, update_ui, passwords)  
    
    # Start the Wi-Fi password retrieval in a background thread
    threading.Thread(target=run_wifi_passwords).start()

# Automatically load Wi-Fi speed and passwords when Tab 3 is loaded
def on_tab_selected(event):
    
    if notebook.index("current") == 2:  
        display_wifi_speed()  
        display_wifi_passwords()  

# Bind tab selection event to the handler
notebook.bind("<<NotebookTabChanged>>", on_tab_selected)


# Tab 4: Remove Temporary Files
temp_files_tab = ttk.Frame(notebook)
notebook.add(temp_files_tab, text="Temp File Cleaner")

# Frame for Display and Actions
temp_frame = ttk.Frame(temp_files_tab)
temp_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

# Label for Information
temp_label = ttk.Label(temp_frame, text="Temporary Files Cleaner", font=("Helvetica", 14, "bold"))
temp_label.pack(anchor=tk.W, pady=5)

# ScrolledText for Displaying Temp Files
temp_display = ScrolledText(temp_frame, height=15, wrap=tk.WORD, font=("Helvetica", 12))
temp_display.pack(fill=tk.BOTH, expand=True, pady=5)

# Function to List Temp Files
def list_temp_files():
    temp_display.delete(1.0, tk.END)  
    temp_dirs = [os.getenv("TEMP"), os.path.join(os.getenv("SYSTEMROOT"), "Temp")]

    temp_files = []
    for temp_dir in temp_dirs:
        if os.path.exists(temp_dir):
            for root, _, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    temp_files.append(file_path)

    if temp_files:
        for file_path in temp_files:
            temp_display.insert(tk.END, f"{file_path}\n")
    else:
        temp_display.insert(tk.END, "No temporary files found.")

# Function to Remove Temp Files
def remove_temp_files():
    temp_files = temp_display.get(1.0, tk.END).strip().split("\n")
    deleted_count = 0

    for file_path in temp_files:
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                deleted_count += 1
        except Exception as e:
            print(f"Error deleting {file_path}: {e}")

    temp_display.delete(1.0, tk.END)
    temp_display.insert(tk.END, f"Deleted {deleted_count} temporary files.")

# Buttons for Actions
ttk.Button(temp_frame, text="Scan Temp Files", command=list_temp_files).pack(side=tk.LEFT, padx=5, pady=10)
ttk.Button(temp_frame, text="Delete Temp Files", command=remove_temp_files).pack(side=tk.LEFT, padx=5, pady=10)



# Tab 5: Actions
actions_tab = ttk.Frame(notebook)
notebook.add(actions_tab, text="Actions")

# Handle removal of temporary files
def handle_remove_temp_files():
    actions_tab.delete(1.0, tk.END)  

# Initialize Real-Time Monitoring
update_graphs()
check_alerts()

# Short Button for Actions Tab
ttk.Button(actions_tab, text="Shutdown", width=20, command=shutdown).pack(pady=5)
ttk.Button(actions_tab, text="Restart", width=20, command=restart).pack(pady=5)

root.mainloop()
